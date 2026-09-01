"""Rakuten Price Comparison and Excel Highlighter.

Compares scraped product prices from Rakuten store Excel files against
the official catalog prices configured in environment by model code.

Color Highlights:
- RED (#FFC7CE): Code exists in official catalog, but price differs.
- YELLOW (#FFF2CC): Product code not found in catalog or unassigned.
- GREEN (#C6EFCE): Price matches official catalog price.
"""

import sys
from typing import Any, Dict, List, Optional, Set, Tuple

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import pandas as pd

from config import (
    CATALOG_LIST_EXCEL,
    OUTPUT_COMPARISON_EXCEL,
    OUTPUT_SCRAPED_EXCEL,
    TARGET_KEYWORD,
)
from utils import (
    evaluate_point_status,
    format_currency_yen,
    parse_numeric_price,
)

# Configure UTF-8 encoding for Windows console output
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")


def load_official_catalog_prices(
    list_products_file: str = CATALOG_LIST_EXCEL,
) -> Dict[str, Dict[str, Optional[float]]]:
    """Load official catalog prices mapped by uppercase product model code.

    Args:
        list_products_file: Path to catalog Excel file.

    Returns:
        Dictionary mapping upper product codes to tax-included and
        tax-excluded prices.
    """
    catalog_prices: Dict[str, Dict[str, Optional[float]]] = {}
    try:
        df_list = pd.read_excel(list_products_file, header=None)
        for idx in range(len(df_list)):
            row_vals = [
                str(v) for v in df_list.iloc[idx].values if pd.notna(v)
            ]
            if any("型番" in v for v in row_vals):
                for r in range(idx + 1, len(df_list)):
                    row_r = df_list.iloc[r]
                    code_val = row_r[1]
                    if pd.notna(code_val):
                        code = str(code_val).strip()
                        if code and code.lower() != "nan" and code != "型番":
                            p_excl = parse_numeric_price(row_r[5])
                            p_incl = parse_numeric_price(row_r[6])
                            catalog_prices[code.upper()] = {
                                "tax_excluded": p_excl,
                                "tax_included": p_incl,
                            }
                break
    except Exception as exc:
        print(
            f"Error loading catalog file '{list_products_file}': {exc}",
            flush=True,
        )

    return catalog_prices


def compare_single_item(
    product_code_raw: Any,
    scraped_price_raw: Any,
    catalog_prices: Dict[str, Dict[str, Optional[float]]],
) -> Tuple[str, Optional[float], Optional[float]]:
    """Compare a single scraped item against official catalog prices.

    Args:
        product_code_raw: Scraped product code.
        scraped_price_raw: Scraped price string.
        catalog_prices: Loaded catalog price map.

    Returns:
        Tuple of (status, official_tax_included_price,
        official_tax_excluded_price).
    """
    p_code = (
        str(product_code_raw).strip().upper()
        if pd.notna(product_code_raw)
        else ""
    )
    scraped_price = parse_numeric_price(scraped_price_raw)

    invalid_codes: Set[str] = {
        "",
        TARGET_KEYWORD.upper(),
        "NAN",
        "NONE",
        "SIN CÓDIGO ESPECÍFICO",
        "SIN CODIGO ESPECIFICO",
    }

    if not p_code or p_code in invalid_codes:
        return "CODE_NOT_FOUND", None, None

    if p_code in catalog_prices:
        info = catalog_prices[p_code]
        p_incl = info["tax_included"]
        p_excl = info["tax_excluded"]

        if scraped_price is not None:
            if p_incl is not None and abs(scraped_price - p_incl) < 1.0:
                return "MATCH", p_incl, p_excl
            if p_excl is not None and abs(scraped_price - p_excl) < 1.0:
                return "MATCH", p_incl, p_excl

        return "PRICE_MISMATCH", p_incl, p_excl

    return "CODE_NOT_FOUND", None, None


def compare_and_highlight_excel(
    scraped_excel_input: str = OUTPUT_SCRAPED_EXCEL,
    list_products_file: str = CATALOG_LIST_EXCEL,
    output_excel: str = OUTPUT_COMPARISON_EXCEL,
    check_points: bool = True,
) -> None:
    """Compare prices and color-highlight Excel rows based on status.

    Highlights:
    - RED (#FFC7CE): Price Mismatch (Code exists, but price differs).
    - YELLOW (#FFF2CC): Code Not Found in catalog.
    - GREEN (#C6EFCE): Price Match.

    Args:
        scraped_excel_input: Input Excel with scraped product data.
        list_products_file: Official catalog Excel file.
        output_excel: Output Excel file with highlights.
        check_points: Whether to evaluate and include the Point Status column.
    """
    catalog_prices = load_official_catalog_prices(list_products_file)
    print(
        f"Loaded {len(catalog_prices)} official product prices for "
        "comparison.",
        flush=True,
    )

    try:
        excel_file = pd.ExcelFile(scraped_excel_input)
    except Exception as exc:
        print(
            f"Could not open input file '{scraped_excel_input}': {exc}",
            flush=True,
        )
        return

    sheet_names = excel_file.sheet_names
    print(
        f"Found {len(sheet_names)} sheet(s) in '{scraped_excel_input}'.",
        flush=True,
    )

    # Step 1: Process and write updated DataFrames with comparison columns
    try:
        with pd.ExcelWriter(output_excel, engine="openpyxl") as writer:
            for s_name in sheet_names:
                df = pd.read_excel(excel_file, sheet_name=s_name)

                statuses: List[str] = []
                point_statuses: List[str] = []
                off_incl_list: List[str] = []
                off_excl_list: List[str] = []

                for _, row in df.iterrows():
                    p_code = row.get("Product_Code")
                    p_price = row.get("Price")
                    p_points = row.get("Points")

                    status, p_incl, p_excl = compare_single_item(
                        p_code, p_price, catalog_prices
                    )

                    statuses.append(status)
                    off_incl_list.append(format_currency_yen(p_incl))
                    off_excl_list.append(format_currency_yen(p_excl))

                    if check_points:
                        point_statuses.append(evaluate_point_status(p_points))

                df["Official_Price_Incl_Tax"] = off_incl_list
                df["Official_Price_Excl_Tax"] = off_excl_list
                if check_points:
                    df["Point Status"] = point_statuses
                df["Comparison_Status"] = statuses

                df.to_excel(writer, sheet_name=str(s_name), index=False)

    except PermissionError:
        fallback = (
            output_excel[:-5] + "_updated.xlsx"
            if output_excel.endswith(".xlsx")
            else output_excel + "_updated.xlsx"
        )
        with pd.ExcelWriter(fallback, engine="openpyxl") as writer:
            for s_name in sheet_names:
                df = pd.read_excel(excel_file, sheet_name=s_name)

                statuses = []
                point_statuses = []
                off_incl_list = []
                off_excl_list = []

                for _, row in df.iterrows():
                    p_code = row.get("Product_Code")
                    p_price = row.get("Price")
                    p_points = row.get("Points")

                    status, p_incl, p_excl = compare_single_item(
                        p_code, p_price, catalog_prices
                    )

                    statuses.append(status)
                    off_incl_list.append(format_currency_yen(p_incl))
                    off_excl_list.append(format_currency_yen(p_excl))

                    if check_points:
                        point_statuses.append(evaluate_point_status(p_points))

                df["Official_Price_Incl_Tax"] = off_incl_list
                df["Official_Price_Excl_Tax"] = off_excl_list
                if check_points:
                    df["Point Status"] = point_statuses
                df["Comparison_Status"] = statuses

                df.to_excel(writer, sheet_name=str(s_name), index=False)
        output_excel = fallback

    print("Data processing complete. Applying color highlights...", flush=True)

    # Step 2: Open generated Excel with openpyxl to apply row fills
    wb = load_workbook(output_excel)

    red_fill = PatternFill(
        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
    )
    yellow_fill = PatternFill(
        start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
    )
    green_fill = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    )

    total_matched = 0
    total_mismatched = 0
    total_not_found = 0
    total_points_ok = 0
    total_points_x = 0

    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        header = [cell.value for cell in ws[1]]

        try:
            status_col_idx = header.index("Comparison_Status") + 1
        except ValueError:
            continue

        point_col_idx = (
            header.index("Point Status") + 1
            if "Point Status" in header
            else None
        )

        for row_idx in range(2, ws.max_row + 1):
            status_val = ws.cell(row=row_idx, column=status_col_idx).value

            target_fill = None
            if status_val == "PRICE_MISMATCH":
                target_fill = red_fill
                total_mismatched += 1
            elif status_val == "CODE_NOT_FOUND":
                target_fill = yellow_fill
                total_not_found += 1
            elif status_val == "MATCH":
                target_fill = green_fill
                total_matched += 1

            if point_col_idx:
                p_stat = ws.cell(row=row_idx, column=point_col_idx).value
                if p_stat == "OK":
                    total_points_ok += 1
                elif p_stat == "X":
                    total_points_x += 1

            if target_fill:
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = target_fill

    wb.save(output_excel)

    print(
        "\n--- Price Comparison Summary ---",
        flush=True,
    )
    print(f"Total Matches (GREEN): {total_matched}", flush=True)
    print(f"Total Mismatches (RED): {total_mismatched}", flush=True)
    print(f"Total Code Not Found (YELLOW): {total_not_found}", flush=True)
    if check_points:
        print(
            f"Point Status: {total_points_ok} OK, {total_points_x} X",
            flush=True,
        )
    print(
        f"Highlighted Excel report generated at '{output_excel}'!",
        flush=True,
    )


if __name__ == "__main__":
    compare_and_highlight_excel(
        scraped_excel_input=OUTPUT_SCRAPED_EXCEL,
        list_products_file=CATALOG_LIST_EXCEL,
        output_excel=OUTPUT_COMPARISON_EXCEL,
        check_points=True,
    )
